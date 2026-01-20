import streamlit as st
import xml.etree.ElementTree as ET
import pandas as pd
from io import BytesIO
from datetime import datetime
import zipfile
import tempfile
import os
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def formatar_excel(writer, sheet_name, df):
    """Formata a planilha Excel com cabeçalho azul e valores em formato contábil"""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Colunas numéricas que devem ter formato contábil
    colunas_valores = ['Qtd_Item', 'Valor_Unit', 'Valor_Prod', 'Desconto',
                       'Aliq_ICMS', 'Valor_ICMS', 'Aliq_PIS', 'Valor_PIS',
                       'Aliq_Cofins', 'Valor_Cofins', 'BC_IBS_CBS',
                       'Aliq_IBS_UF', 'Valor_IBS_UF', 'Aliq_IBS_Mun', 'Valor_IBS_Mun',
                       'Aliq_CBS', 'Valor_CBS', 'Qtd_Ocorrencias']
    
    # Formatar cabeçalho (primeira linha) com fundo azul e texto branco
    header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Formatar colunas de valores com formato contábil
    for col_num, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)
        
        # Ajustar largura da coluna
        max_length = max(len(str(col_name)), 15)
        worksheet.column_dimensions[col_letter].width = max_length + 2
        
        # Aplicar formato contábil para colunas de valores
        if col_name in colunas_valores:
            for row_num in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                
                # Converter valor para numérico se necessário
                valor_original = df.iloc[row_num - 2][col_name]
                if pd.notna(valor_original) and valor_original != '':
                    try:
                        valor_numerico = float(valor_original)
                        cell.value = valor_numerico
                    except (ValueError, TypeError):
                        pass
                
                # Formato contábil brasileiro
                if col_name == 'Qtd_Item' or col_name == 'Qtd_Ocorrencias':
                    cell.number_format = '#,##0.00'  # Formato numérico com 2 casas
                elif 'Aliq' in col_name:
                    cell.number_format = '0.00'  # Formato numérico simples para alíquotas
                else:
                    # Formato Contábil R$ (usa _ para alinhar valores)
                    cell.number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

def extrair_dados_fiscais(arquivo_xml):
    tree = ET.parse(arquivo_xml)
    root = tree.getroot()

    ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

    dados_list = []

    for nfe in root.findall('.//nfe:NFe', ns):
        dados = {}
        infNFe = nfe.find('nfe:infNFe', ns)
        if infNFe is not None:
            ide = infNFe.find('nfe:ide', ns)
            if ide is not None:
                dados['Num_Doc'] = ide.findtext('nfe:nNF', '', ns)
                dados['Serie'] = ide.findtext('nfe:serie', '', ns)
                
                # Formatar data para DD/MM/AAAA
                data_emissao = ide.findtext('nfe:dhEmi', '', ns)
                if data_emissao:
                    try:
                        # Formato típico: 2024-01-19T10:30:00-03:00
                        dt = datetime.fromisoformat(data_emissao.replace('Z', '+00:00'))
                        dados['Data_Emissao'] = dt.strftime('%d/%m/%Y')
                    except:
                        dados['Data_Emissao'] = data_emissao
                else:
                    dados['Data_Emissao'] = ''

            emit = infNFe.find('nfe:emit', ns)
            if emit is not None:
                dados['CNPJ'] = emit.findtext('nfe:CNPJ', '', ns)

            for det in infNFe.findall('nfe:det', ns):
                det_dados = dados.copy()
                prod = det.find('nfe:prod', ns)
                if prod is not None:
                    det_dados['Codigo_Prod'] = prod.findtext('nfe:cProd', '', ns)
                    det_dados['Descricao_Produto'] = prod.findtext('nfe:xProd', '', ns)
                    det_dados['NCM'] = prod.findtext('nfe:NCM', '', ns)
                    det_dados['CFOP'] = prod.findtext('nfe:CFOP', '', ns)
                    det_dados['Qtd_Item'] = prod.findtext('nfe:qCom', '', ns)
                    det_dados['Valor_Unit'] = prod.findtext('nfe:vUnCom', '', ns)
                    det_dados['Valor_Prod'] = prod.findtext('nfe:vProd', '', ns)
                    det_dados['Desconto'] = prod.findtext('nfe:vDesc', '', ns)

                imposto = det.find('nfe:imposto', ns)
                if imposto is not None:
                    # ICMS - buscar em diferentes grupos
                    icms_node = None
                    for icms_type in ['ICMS00', 'ICMS10', 'ICMS20', 'ICMS30', 'ICMS40', 'ICMS51', 'ICMS60', 'ICMS70', 'ICMS90', 'ICMSSN101', 'ICMSSN102', 'ICMSSN201', 'ICMSSN202', 'ICMSSN500', 'ICMSSN900']:
                        icms_node = imposto.find(f'.//nfe:{icms_type}', ns)
                        if icms_node is not None:
                            break
                    
                    if icms_node is not None:
                        det_dados['ICMS_CST'] = icms_node.findtext('nfe:CST', '', ns) or icms_node.findtext('nfe:CSOSN', '', ns)
                        det_dados['Aliq_ICMS'] = icms_node.findtext('nfe:pICMS', '', ns)
                        det_dados['Valor_ICMS'] = icms_node.findtext('nfe:vICMS', '', ns)
                    
                    # PIS - buscar em diferentes grupos
                    pis_node = None
                    for pis_type in ['PISAliq', 'PISQtde', 'PISNT', 'PISOutr']:
                        pis_node = imposto.find(f'.//nfe:{pis_type}', ns)
                        if pis_node is not None:
                            break
                    
                    if pis_node is not None:
                        det_dados['CST_PIS'] = pis_node.findtext('nfe:CST', '', ns)
                        det_dados['Aliq_PIS'] = pis_node.findtext('nfe:pPIS', '', ns)
                        det_dados['Valor_PIS'] = pis_node.findtext('nfe:vPIS', '', ns)

                    # COFINS - buscar em diferentes grupos
                    cofins_node = None
                    for cofins_type in ['COFINSAliq', 'COFINSQtde', 'COFINSNT', 'COFINSOutr']:
                        cofins_node = imposto.find(f'.//nfe:{cofins_type}', ns)
                        if cofins_node is not None:
                            break
                    
                    if cofins_node is not None:
                        det_dados['CST_Cofins'] = cofins_node.findtext('nfe:CST', '', ns)
                        det_dados['Aliq_Cofins'] = cofins_node.findtext('nfe:pCOFINS', '', ns)
                        det_dados['Valor_Cofins'] = cofins_node.findtext('nfe:vCOFINS', '', ns)
                    
                    # IBS/CBS
                    ibscbs = imposto.find('.//nfe:IBSCBS', ns)
                    if ibscbs is not None:
                        det_dados['CST_IBS'] = ibscbs.findtext('nfe:CST', '', ns)
                        det_dados['cClassTrib'] = ibscbs.findtext('nfe:cClassTrib', '', ns)
                        
                        # Base de cálculo
                        det_dados['BC_IBS_CBS'] = ibscbs.findtext('.//nfe:vBC', '', ns)
                        
                        # IBS UF
                        det_dados['Aliq_IBS_UF'] = ibscbs.findtext('.//nfe:pIBSUF', '', ns)
                        det_dados['Valor_IBS_UF'] = ibscbs.findtext('.//nfe:vIBSUF', '', ns)
                        
                        # IBS Municipal
                        det_dados['Aliq_IBS_Mun'] = ibscbs.findtext('.//nfe:pIBSMun', '', ns)
                        det_dados['Valor_IBS_Mun'] = ibscbs.findtext('.//nfe:vIBSMun', '', ns)
                        
                        # CBS
                        det_dados['Aliq_CBS'] = ibscbs.findtext('.//nfe:pCBS', '', ns)
                        det_dados['Valor_CBS'] = ibscbs.findtext('.//nfe:vCBS', '', ns)
                
                dados_list.append(det_dados)

    return dados_list

st.title('Extrator de Dados Fiscais de XML')

uploaded_file = st.file_uploader("Escolha um arquivo XML ou ZIP", type=["xml", "zip"])

if uploaded_file is not None:
    todos_dados_fiscais = []
    
    # Verificar se é um arquivo ZIP
    if uploaded_file.name.endswith('.zip'):
        with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
            # Listar arquivos XML dentro do ZIP
            xml_files = [f for f in zip_ref.namelist() if f.endswith('.xml')]
            
            if not xml_files:
                st.error("Nenhum arquivo XML encontrado no ZIP.")
            else:
                st.info(f"Encontrados {len(xml_files)} arquivo(s) XML no ZIP.")
                
                # Processar cada arquivo XML
                for xml_file in xml_files:
                    try:
                        with zip_ref.open(xml_file) as xml_content:
                            dados = extrair_dados_fiscais(xml_content)
                            todos_dados_fiscais.extend(dados)
                    except Exception as e:
                        st.warning(f"Erro ao processar {xml_file}: {str(e)}")
    else:
        # Processar arquivo XML único
        dados_fiscais = extrair_dados_fiscais(uploaded_file)
        todos_dados_fiscais.extend(dados_fiscais)
    
    if todos_dados_fiscais:
        st.success(f"Dados extraídos com sucesso! Total de {len(todos_dados_fiscais)} registro(s).")
        
        df = pd.DataFrame(todos_dados_fiscais)
        
        # Criar abas
        tab1, tab2 = st.tabs(["📋 Dados Completos", "📊 Resumo por Item"])
        
        with tab1:
            st.subheader("Dados Completos Extraídos")
            
            # Definir ordem das colunas
            colunas_ordenadas = [
                'Num_Doc', 'Serie', 'Data_Emissao', 'CNPJ', 'Codigo_Prod', 'Descricao_Produto',
                'NCM', 'CFOP', 'Qtd_Item', 'Valor_Unit', 'Valor_Prod', 'Desconto',
                'ICMS_CST', 'Aliq_ICMS', 'Valor_ICMS', 
                'CST_PIS', 'Aliq_PIS', 'Valor_PIS', 
                'CST_Cofins', 'Aliq_Cofins', 'Valor_Cofins',
                'CST_IBS', 'cClassTrib', 'BC_IBS_CBS', 
                'Aliq_IBS_UF', 'Valor_IBS_UF', 'Aliq_IBS_Mun', 'Valor_IBS_Mun', 
                'Aliq_CBS', 'Valor_CBS'
            ]
            # Manter apenas colunas que existem no DataFrame
            colunas_finais = [col for col in colunas_ordenadas if col in df.columns]
            df = df[colunas_finais]
            
            st.dataframe(df)
            
            # Botão de download para dados completos
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Dados Fiscais')
                formatar_excel(writer, 'Dados Fiscais', df)
            
            st.download_button(
                label="📥 Baixar dados completos em Excel",
                data=output.getvalue(),
                file_name='dados_fiscais_completo.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        with tab2:
            st.subheader("Resumo por Item")
            
            # Criar resumo agrupado por item
            colunas_resumo = ['Descricao_Produto', 'CST_IBS', 'cClassTrib']
            colunas_existentes = [col for col in colunas_resumo if col in df.columns]
            
            if colunas_existentes:
                # Adicionar colunas numéricas para resumo
                colunas_numericas = ['Qtd_Item', 'Valor_Unit', 'Valor_Prod', 'Desconto', 
                                    'Aliq_ICMS', 'Valor_ICMS', 
                                    'Aliq_PIS', 'Valor_PIS', 
                                    'Aliq_Cofins', 'Valor_Cofins',
                                    'BC_IBS_CBS', 'Aliq_IBS_UF', 'Valor_IBS_UF', 
                                    'Aliq_IBS_Mun', 'Valor_IBS_Mun', 'Aliq_CBS', 'Valor_CBS']
                
                # Converter colunas numéricas
                for col in colunas_numericas:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce')
                
                # Agrupar por produto e CST
                if 'Descricao_Produto' in df.columns:
                    grupo_cols = ['Descricao_Produto']
                    if 'CST_IBS' in df.columns:
                        grupo_cols.append('CST_IBS')
                    if 'cClassTrib' in df.columns:
                        grupo_cols.append('cClassTrib')
                    
                    # Criar agregações
                    agg_dict = {}
                    for col in colunas_numericas:
                        if col in df.columns:
                            agg_dict[col] = 'sum'
                    
                    if agg_dict:
                        df_resumo = df.groupby(grupo_cols, dropna=False).agg(agg_dict).reset_index()
                        
                        # Adicionar contagem de ocorrências
                        df_resumo.insert(1, 'Qtd_Ocorrencias', df.groupby(grupo_cols, dropna=False).size().values)
                        
                        st.dataframe(df_resumo)
                        
                        # Botão de download para resumo
                        output_resumo = BytesIO()
                        with pd.ExcelWriter(output_resumo, engine='openpyxl') as writer:
                            df_resumo.to_excel(writer, index=False, sheet_name='Resumo por Item')
                            formatar_excel(writer, 'Resumo por Item', df_resumo)
                        
                        st.download_button(
                            label="📥 Baixar resumo em Excel",
                            data=output_resumo.getvalue(),
                            file_name='resumo_por_item.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                    else:
                        st.warning("Nenhuma coluna numérica encontrada para resumir.")
                else:
                    st.warning("Coluna 'Descricao_Produto' não encontrada nos dados.")
            else:
                st.warning("Colunas necessárias para o resumo não foram encontradas nos dados extraídos.")
    else:
        st.error("Não foi possível extrair dados do arquivo XML. Verifique o formato do arquivo.")